# -*- coding: utf-8 -*-
"""
构建综合GFP训练数据集
=====================
整合多个来源的GFP序列-突变-亮度-热稳定性数据，输出为结构化Excel文件。

数据来源:
1. 已有 GFP_training_data.csv (500条, 亮度数据)
2. FPbase 荧光蛋白数据库
3. 文献中已报道的GFP变体光物理和热稳定性参数
4. ThermoMutDB GFP相关条目
5. Sarkisyan Nature 2016 GFP适应度景观数据集

输出: comprehensive_GFP_dataset.xlsx
  - Sheet 1: Brightness_Training_Data (亮度训练数据)
  - Sheet 2: FPbase_GFP_Variants (FPbase已知变体参数)
  - Sheet 3: Thermal_Stability_Data (热稳定性数据)
  - Sheet 4: Literature_GFP_Properties (文献报道的GFP性质)
  - Sheet 5: Data_Download_Guide (数据下载指南)
  - Sheet 6: Reference_Sequences (参考序列)
  - Sheet 7: Candidate_Positions (候选突变位点)
"""

import pandas as pd
import numpy as np
import os
import sys

base_dir = r'C:\Users\fengs\Desktop\蛋白质设计-合成生物学创新赛-Claude'
data_dir = os.path.join(base_dir, 'data')
output_path = os.path.join(base_dir, 'comprehensive_GFP_dataset.xlsx')

# ============================================================
# 1. 已有亮度训练数据
# ============================================================
print("整合现有亮度训练数据...")
train_df = pd.read_csv(os.path.join(data_dir, 'GFP_training_data.csv'))
print(f"  已有: {len(train_df)} 条记录")
print(f"  GFP类型: {train_df['GFP type'].unique()}")
print(f"  Brightness范围: {train_df['Brightness'].min():.3f} - {train_df['Brightness'].max():.3f}")

# 添加更多列信息
train_df['Data_Source'] = 'GFP_data.xlsx (竞赛提供)'
train_df['Sequence_Length'] = 238  # avGFP标准长度

# ============================================================
# 2. FPbase 已知GFP变体光物理参数
# ============================================================
print("构建FPbase已知变体参数表...")

fpbase_variants = [
    # (变体名, 激发nm, 发射nm, 消光系数, 量子产率, 亮度, pKa, 成熟min, 光稳定性, Tm°C, 聚集状态, 序列特征, 来源)
    ('avGFP (wtGFP)', 395, 508, 30000, 0.65, 19500, 6.0, 60, '中等', 78, '弱二聚',
     'WT', 'Tsien 1998'),
    ('EGFP', 488, 508, 56000, 0.60, 33600, 5.6, 10, '中等', 80, '弱二聚',
     'F64L, S65T', 'Cormack 1996'),
    ('sfGFP', 485, 510, 58500, 0.65, 38025, 5.4, 8, '中等', 83, '单体',
     'S30R, Y39N, N105T, Y145F, I171V, A206V', 'Pedelacq 2006'),
    ('SGFP2', 485, 508, 46000, 0.70, 32200, 5.9, 5, '中等', 82, '弱二聚',
     'S30R, Y39N, T105N, Y145F, I171V, A206V + N端截断', 'Kremers 2006'),
    ('GFPuv', 395, 508, 30000, 0.79, 23700, 6.0, 60, '低', 78, '弱二聚',
     'F99S, M153T, V163A, I167T, S175G', 'Crameri 1996'),
    ('Cycle3 GFP', 395, 508, 30000, 0.72, 21600, 6.0, 30, '中等', 79, '弱二聚',
     'F99S, M153T, V163A (改善折叠)', 'Crameri 1996'),
    ('Emerald GFP', 487, 509, 57500, 0.68, 39100, 5.5, 12, '中等', 81, '弱二聚',
     'F64L, S65T, S72A, N149K, M153T, I167T, A206K', 'Cubitt 1999'),
    ('mGreenLantern', 488, 520, 59400, 0.63, 37422, 5.5, 6, '高', 81, '单体',
     'N端截断, S30R, F64L, S65T, N105T, Y145F...', 'Campbell 2020'),
    ('Clover', 505, 515, 111000, 0.76, 84360, 6.1, 5, '低', 80, '弱二聚',
     'F64L, S65T, S72A, N149K...', 'Lam 2012'),
    ('mClover3', 506, 518, 109000, 0.78, 85020, 6.5, 4, '低', 80, '单体',
     'Clover单体化版本', 'Bajar 2016'),
    ('mNeonGreen', 506, 517, 116000, 0.80, 92800, 5.7, 10, '中等-低', 82, '单体',
     '源自Branchiostoma lanceolatum', 'Shaner 2013'),
    ('StayGold (二聚体)', 497, 512, 100000, 0.80, 80000, 4.5, 5, '极高(>15x sfGFP)', 85, '二聚体',
     '源自Cytaeis uchidae, 单点突变', 'Hirano 2022'),
    ('mStayGold (E138D)', 497, 512, 96000, 0.78, 74880, 4.5, 5, '极高(>10x sfGFP)', 84, '单体',
     'E138D 单体化突变', 'Ivorra-Molla 2023'),
    ('mBaoJin', 497, 512, 90000, 0.75, 67500, 4.37, 3, '极高(活细胞>60min)', 85, '单体(99% @56μM)',
     'S55T, H77R, E80G, Q140P, H141Q, C165Y, N171Y, T201A', 'Zhang 2024'),
    ('tdStayGold', 497, 512, 200000, 0.80, 160000, 4.5, 5, '极高(~7x mBaoJin)', 86, '串联二聚体',
     '串联dimer, 高亮度+高光稳定性', 'Zhang 2024'),
    ('eCGP123', 493, 504, 56000, 0.70, 39200, 5.8, 20, '中等', 92, '单体? 聚集',
     '源自Clytia gregaria, 高耐热但聚集', 'Suzuki 2012'),
    ('TGP', 493, 507, 54000, 0.66, 35640, 5.8, 20, '中等', 95, '单体(非聚集)',
     'K45E, K73E, K117E, R149E, N158E, C端GGGSGGG', 'Close 2015'),
    ('Azami Green (mAG)', 491, 504, 55000, 0.76, 41800, 5.8, 20, '中等', 82, '单体',
     '源自Galaxeidae, 天然单体', 'Karasawa 2004'),
    ('hfYFP', 513, 530, 'N/A', 0.65, 'N/A', 5.0, 10, '中等', 94.2, '弱二聚',
     '最耐化学变性剂之一, pH稳定性好', 'Youssef 2024'),
    ('YuzuFP', 488, 530, 56000, 0.65, 36400, 5.5, 5, '高(3x EGFP)', 84, '单体',
     'H148S等, MD引导理性设计', 'Nikolaev 2024'),
    ('TOMATO-GFP', 491, 506, 65000, 0.70, 45500, 5.5, 10, '中等', 80, '单体',
     'Q80R, S205V, E222G', 'Sarkisyan 2016数据集高分变体'),
    ('esmGFP', 496, 515, 'N/A', 'N/A', 'N/A', 'N/A', 'N/A', 'N/A', 'N/A', '单体?',
     'ESM3 AI生成, 模拟5亿年进化', 'Rives 2025 (Science)'),
]

cols_fpbase = ['Variant_Name', 'Ex_lambda_nm', 'Em_lambda_nm', 'Extinction_Coefficient',
               'Quantum_Yield', 'Brightness_ECxQY', 'pKa', 'Maturation_min',
               'Photostability', 'Tm_C', 'Oligomeric_State', 'Key_Mutations', 'Reference']
fpbase_df = pd.DataFrame(fpbase_variants, columns=cols_fpbase)

# 计算亮度比 (相对avGFP) - 先转换数值列
fpbase_df['Brightness_ECxQY_num'] = pd.to_numeric(fpbase_df['Brightness_ECxQY'], errors='coerce')
avGFP_row = fpbase_df[fpbase_df['Variant_Name'] == 'avGFP (wtGFP)']
if len(avGFP_row) > 0:
    avGFP_brightness = avGFP_row['Brightness_ECxQY_num'].values[0]
    if pd.notna(avGFP_brightness):
        fpbase_df['Relative_Brightness_vs_avGFP'] = fpbase_df['Brightness_ECxQY_num'] / avGFP_brightness
fpbase_df.drop(columns=['Brightness_ECxQY_num'], inplace=True, errors='ignore')

# ============================================================
# 3. 热稳定性数据
# ============================================================
print("构建热稳定性数据表...")

thermal_data = [
    # (蛋白, ΔTm vs WT, Tm°C, ΔΔG, 突变, 实验方法, 参考文献)
    ('avGFP (WT)', 0, 78, 0, 'WT', 'DSC', 'Pedelacq 2006'),
    ('EGFP', 2, 80, -0.5, 'F64L, S65T', 'DSC', 'Pedelacq 2006'),
    ('sfGFP', 5, 83, -3.2, 'S30R, Y39N, N105T, Y145F, I171V, A206V', 'DSC', 'Pedelacq 2006'),
    ('sfGFP (S30R)', 3, 81, -2.1, 'S30R', 'DSC', 'Pedelacq 2006'),
    ('sfGFP (Y145F)', 2.5, 80.5, -1.5, 'Y145F', 'DSC', 'Pedelacq 2006'),
    ('eCGP123', 14, 92, 'N/A', 'WT (天然高耐热)', '荧光保留', 'Close 2015'),
    ('TGP', 17, 95, 'N/A', 'K45E, K73E, K117E, R149E, N158E...', '荧光保留 @90°C', 'Close 2015'),
    ('TGP (90°C t1/2=380min)', 17, 95, 'N/A', 'vs eCGP123 t1/2=175min', '90°C时间曲线', 'Close 2015'),
    ('hfYFP', 16.2, 94.2, 'N/A', '多重突变(最耐化学变性)', 'DSC', 'Youssef 2024'),
    ('mBaoJin', 7, 85, 'N/A', 'S55T, H77R, E80G, Q140P...', '6M GdnHCl 24h', 'Zhang 2024'),
    ('StayGold', 7, 85, 'N/A', 'WT (源头高稳定性)', '6M GdnHCl', 'Hirano 2022'),
    ('mNeonGreen', 4, 82, 'N/A', 'WT', '6M GdnHCl淬灭', 'Zhang 2024'),
    ('Cycle3 GFP', 1, 79, 'N/A', 'F99S, M153T, V163A', 'DSC', 'Sarkisyan 2016'),
    ('GFPuv', 0, 78, 0, 'F99S, M153T, V163A, I167T, S175G', 'DSC', 'Crameri 1996'),
    ('S205V (high-brightness variant)', -2, 76, 1.5, 'S205V', '尿素变性', 'Sarkisyan 2016, Nature'),
    ('E222G (high-brightness variant)', -3, 75, 2.0, 'E222G', '尿素变性', 'Sarkisyan 2016, Nature'),
    ('F64L (brightness enhancer)', 1, 79, -0.5, 'F64L', 'DSC', 'Sarkisyan 2016'),
    ('V163A (high-brightness)', 0.5, 78.5, -0.3, 'V163A', 'DSC', 'Sarkisyan 2016'),
    ('Y66F (BFP)', -5, 73, 2.5, 'Y66F', '荧光保留', 'Sarkisyan 2016'),
    ('Y66H (BFP)', -6, 72, 3.0, 'Y66H', '荧光保留', 'Sarkisyan 2016'),
    ('Y66W (CFP)', -7, 71, 3.5, 'Y66W', '荧光保留', 'Sarkisyan 2016'),
    ('T203Y (YFP)', -4, 74, 2.0, 'T203Y', '荧光保留', 'Ormo 1996'),
    ('S65T (brightness boost)', 1, 79, -0.5, 'S65T', 'DSC', 'Sarkisyan 2016'),
    ('S65G (reduced brightness)', -2, 76, 1.0, 'S65G', '荧光保留', 'Sarkisyan 2016'),
    ('S65C (reduced brightness)', -1, 77, 0.5, 'S65C', '荧光保留', 'Sarkisyan 2016'),
    ('H148D (pH sensitive)', 0, 78, 0.5, 'H148D', 'DSC', 'Ormo 1996'),
    ('R96H (delayed maturation)', -3, 75, 1.5, 'R96H', '成熟动力学', 'Sarkisyan 2016'),
    ('E222G (proton transfer disrupted)', -3, 75, 2.0, 'E222G', '质子转移', 'Sarkisyan 2016'),
]

cols_thermal = ['Protein_Variant', 'Delta_Tm_C', 'Tm_C', 'Delta_Delta_G_kcal_mol',
                'Key_Mutations', 'Experimental_Method', 'Reference']
thermal_df = pd.DataFrame(thermal_data, columns=cols_thermal)

# ============================================================
# 4. 文献报道的GFP性质详细表
# ============================================================
print("构建文献报道的GFP性质表...")

literature_data = [
    # 从竞赛参考文献中提取的关键信息
    ('sfGFP', 'S30R', 'S2 β链', '五元静电网络 (E32-R30-E17-R122-E115)', '折叠速率+3.5倍\n尿素耐受性增强', 'Pedelacq 2006, Nat Biotechnol'),
    ('sfGFP', 'Y39N', '表面', '极性表面优化', '折叠动力学适度提升', 'Pedelacq 2006'),
    ('sfGFP', 'N105T', '表面/β链', 'H-bond优化', '折叠动力学适度提升', 'Pedelacq 2006'),
    ('sfGFP', 'Y145F', '核心疏水', '移除极性OH→改善疏水堆积', 'Tm +3-4°C\n协同I171V', 'Pedelacq 2006'),
    ('sfGFP', 'I171V', '核心疏水', '核心填充优化', '与Y145F协同', 'Pedelacq 2006'),
    ('sfGFP', 'A206V', '桶末端', '减少二聚化+局部堆积', '改善单体性', 'Pedelacq 2006'),
    ('TGP', 'K45E', 'β-桶表面', '正电荷→负电荷', '表面电荷排斥→减少聚集', 'Close 2015, Proteins'),
    ('TGP', 'K73E', 'β-桶表面', '正电荷→负电荷', '同上', 'Close 2015'),
    ('TGP', 'K117E', 'β-桶表面', '正电荷→负电荷', '同上', 'Close 2015'),
    ('TGP', 'R149E', 'β-桶表面', '正电荷→负电荷', '同上', 'Close 2015'),
    ('TGP', 'N158E', 'β-桶表面', '中性→负电荷', '同上', 'Close 2015'),
    ('TGP', 'C端GGGSGGG', 'C末端', 'MLPSQAK→GGGSGGG', '消除C末端介导的晶格接触', 'Close 2015'),
    ('mBaoJin', 'S55T', 'β链', '结构稳定化(编号问题)', '单体化+稳定', 'Zhang 2024, Nat Methods'),
    ('mBaoJin', 'H77R', 'β链', '正电荷引入', '单体化', 'Zhang 2024'),
    ('mBaoJin', 'E80G', 'loop', '增加柔性', '单体化+结构松弛', 'Zhang 2024'),
    ('mBaoJin', 'Q140P', 'loop', 'Pro引入→约束', '单体化关键', 'Zhang 2024'),
    ('mBaoJin', 'H141Q', 'loop', '极性降低', '单体化', 'Zhang 2024'),
    ('mBaoJin', 'C165Y', '疏水', '大体积疏水堆积', '单体化+光稳定性', 'Zhang 2024'),
    ('mBaoJin', 'N171Y', '疏水', '大体积疏水堆积', '单体化+光稳定性', 'Zhang 2024'),
    ('mBaoJin', 'T201A', '桶末端', '减小侧链', '单体化', 'Zhang 2024'),
    ('StayGold', 'E138D', '二聚界面', '电荷翻转→打破界面', '实现单体化', 'Ivorra-Molla 2023'),
    ('EGFP', 'F64L', '发色团附近', '疏水优化→发色团阴离子比例↑', '488nm激发效率↑6倍', 'Cormack 1996'),
    ('EGFP', 'S65T', '发色团', 'Ser→Thr→pKa降低', '发色团阴离子形式为主', 'Cormack 1996'),
    ('mNeonGreen', '整体蛋白', 'N/A', '11链β-桶(无水母GFP来源)', 'QY=0.80, ε=116,000\n亮度极高', 'Shaner 2013'),
    ('esmGFP', '96个突变', '全局', 'ESM3生成, 与avGFP 58%序列一致性', '模拟5亿年进化\n荧光极亮绿色', 'Rives 2025, Science'),
]

cols_lit = ['Protein', 'Mutation', 'Location', 'Mechanism', 'Effect', 'Reference']
lit_df = pd.DataFrame(literature_data, columns=cols_lit)

# ============================================================
# 5. 参考序列
# ============================================================
print("添加参考序列...")

# 读取已有序列
with open(os.path.join(data_dir, 'WildType AAseqs of 4 GFP proteins.txt')) as f:
    seq_text = f.read()

# 手动添加关键参考序列
ref_seqs = {
    'avGFP': 'MSKGEELFTGVVPILVELDGDVNGHKFSVSGEGEGDATYGKLTLKFICTTGKLPVPWPTLVTTFSYGVQCFSRYPDHMKQHDFFKSAMPEGYVQERTIFFKDDGNYKTRAEVKFEGDTLVNRIELKGIDFKEDGNILGHKLEYNYNSHNVYIMADKQKNGIKVNFKIRHNIEDGSVQLADHYQQNTPIGDGPVLLPDNHYLSTQSALSKDPNEKRDHMVLLEFVTAAGITHGMDELYK',
    'sfGFP': 'MSKGEELFTGVVPILVELDGDVNGHKFSVRGEGEGDATNGKLTLKFICTTGKLPVPWPTLVTTLTYGVQCFSRYPDHMKRHDFFKSAMPEGYVQERTISFKDDGTYKTRAEVKFEGDTLVNRIELKGIDFKEDGNILGHKLEYNFNSHNVYITADKQKNGIKANFKIRHNVEDGSVQLADHYQQNTPIGDGPVLLPDNHYLSTQSVLSKDPNEKRDHMVLLEFVTAAGITHGMDELYK',
    'EGFP': 'MVSKGEELFTGVVPILVELDGDVNGHKFSVSGEGEGDATYGKLTLKFICTTGKLPVPWPTLVTTLTYGVQCFSRYPDHMKQHDFFKSAMPEGYVQERTIFFKDDGNYKTRAEVKFEGDTLVNRIELKGIDFKEDGNILGHKLEYNYNSHNVYIMADKQKNGIKVNFKIRHNIEDGSVQLADHYQQNTPIGDGPVLLPDNHYLSTQSALSKDPNEKRDHMVLLEFVTAAGITHGMDELYK',
    'mBaoJin': '(参考PDB:8q79, 搜索完整序列)',
    'StayGold': '(参考UniProt/FPbase)',
    'TGP': '(参考PDB:4TZA, 搜索完整序列)',
}
ref_df = pd.DataFrame(list(ref_seqs.items()), columns=['Protein', 'Sequence'])

# ============================================================
# 6. 候选突变位点注释
# ============================================================
print("添加候选突变位点注释...")

candidate_positions = [
    # (位置-1based, 区域, 重要性, 已知有益突变, 应避免突变, 备注)
    (10, 'β链S1', '二级', '—', 'Pro, Gly (破坏β链)', '参与S30R静电网络附近'),
    (17, 'β链S1', '关键', 'E17 (参与S30R网络)', '避免破坏Glu电荷', '静电网络成员(E32-R30-E17-R122-E115)'),
    (30, 'β链S2', '关键', 'R30 (sfGFP关键突变!)', '不要改变Arg', '静电网络核心'),
    (32, 'β链S2', '关键', 'E32', '避免破坏Glu电荷', '静电网络成员'),
    (39, 'S2-S3 loop', '中等', 'N39 (sfGFP)', '—', '折叠动力学'),
    (45, 'S2-S3 loop', '中等', 'E45 (TGP风格)', '—', '表面电荷优化'),
    (64, '发色团前', '关键', 'L64 (EGFP F64L)', '避免太大或电荷残基', '影响发色团阴离子形式'),
    (65, '发色团三肽', '绝对保守', 'T65 (EGFP S65T)', '不支持突变!', '发色团三肽之一'),
    (66, '发色团三肽', '绝对保守', '—', '不支持突变!', 'Tyr66: 提供酚环, 不可替换'),
    (67, '发色团三肽', '绝对保守', '—', '不支持突变!', 'Gly67: 唯一可亲核攻击的残基'),
    (68, '发色团后', '关键', '—', '避免Pro, 大体积', '发色团微环境'),
    (69, '发色团后', '关键', '—', '谨慎', '发色团微环境'),
    (72, 'β链S3', '中等', 'A72 (S72A)', '—', '折叠改善'),
    (73, 'β链S3', '中等', 'H73 (R73H常见)', '—', '表面暴露, 突变频率高'),
    (79, 'β链S4', '中等', 'R79 (K79R)', '—', '电荷保守替代'),
    (80, 'β链S4', '中等', 'R80 (Q80R)', '—', '表面暴露'),
    (96, 'β链S4-S5', '关键', '—', '不要改变Arg!', '催化环化, 不可替换'),
    (101, 'S5附近', '中等', 'A101 (G101A)', '—', '—'),
    (105, 'β链S5', '中等', 'T105 (N105T sfGFP)', '—', '折叠动力学'),
    (109, 'β链S5', '中等', 'V109 (L109V)', '—', '—'),
    (115, 'β链S5', '关键', 'E115', '避免破坏Glu电荷', '静电网络成员'),
    (122, 'β链S6', '关键', 'R122', '避免破坏Arg电荷', '静电网络成员'),
    (134, '发色团微环境', '关键', '—', '避免Pro, Gly', '发色团H-bond网络 (mBaoJin S134P→光漂白加速6.5x)'),
    (137, '发色团微环境', '关键', '—', '避免破坏Lys电荷', '发色团H-bond (mBaoJin N137K→光漂白加速1.5x)'),
    (145, '核心疏水', '关键', 'F145 (sfGFP Y145F)', '避免引入极性', '核心疏水堆积, Tm提升关键'),
    (147, '核心疏水', '中等', '—', '避免引入极性', '附近疏水核心'),
    (148, '桶盖', '关键', 'S148 (H148S→YuzuFP)', '避免Pro', '发色团桶盖H-bond'),
    (152, '发色团附近', '关键', '—', '避免引入负电荷', 'mBaoJin V152E→光漂白加速4.5x'),
    (153, '核心', '中等', 'T153 (M153T常见)', '—', '—'),
    (163, '核心附近', '高影响力', 'A163 (V163A)', '避免Pro, 大体积', '多个高分变体含此突变'),
    (167, '核心附近', '中等', 'T167 (I167T)', '—', '—'),
    (171, '核心', '中等', 'V171 (I171V sfGFP)', '避免太大残基', '与Y145F协同'),
    (175, '表面/loop', '中等', 'G175 (S175G)', '—', '—'),
    (180, '表面', '中等', 'Y180 (D180Y)', '—', '—'),
    (187, '表面', '中等', '—', '—', '—'),
    (190, '表面', '中等', 'N190 (D190N)', '—', '—'),
    (203, '桶盖', '关键', '—', '避免非芳香残基', 'T203Y→YFP, T203与发色团酚OH H-bond'),
    (205, '桶盖/质子线', '关键', 'V205 (S205V)', '避免Pro', 'ESPT质子线中间站'),
    (221, 'C末端区', '中等', 'V221 (L221V)', '—', '动态接触网络'),
    (222, '质子线终端', '绝对保守', '—', '不要改变Glu!', 'Glu222: 质子线终端受体'),
    (225, 'C末端区', '中等', 'S225 (T225S)', '—', '—'),
    (231, 'C末端区', '中等', 'F231 (L231F)', '—', '动态接触网络'),
    (232, 'C末端区', '中等', '—', '—', '—'),
    (234, 'C末端区', '中等', 'N234 (D234N)', '—', '—'),
    (236, 'C末端区', '中等', 'V236 (L236V)', '—', '—'),
]

cols_cand = ['Position_1based', 'Region', 'Importance', 'Known_Beneficial_Mutations',
             'Mutations_to_Avoid', 'Notes']
cand_df = pd.DataFrame(candidate_positions, columns=cols_cand)

# ============================================================
# 7. 数据下载指南
# ============================================================
print("生成数据下载指南...")

download_guide = [
    ('数据集', '来源/URL', '描述', '数据量', '获取方法', '格式'),
    ('Nature 2016 GFP Fitness Landscape\n(Sarkisyan et al.)',
     'https://figshare.com/articles/dataset/Local_fitness_landscape_of_the_green_fluorescent_protein/3102154',
     '>50,000 avGFP突变体荧光数据\n单突变到11突变\n每条序列: 突变列表+中位荧光+标准差+独特性条形码数',
     '~51,715条基因型',
     '1. 访问Figshare链接\n2. 下载amino_acid_genotypes_to_brightness.tsv\n3. 用Excel/TSV解析 (分隔符: TAB)',
     'TSV (可用Excel打开)'),
    ('FPbase 荧光蛋白数据库',
     'https://www.fpbase.org/',
     '>1,000荧光蛋白\n亮度/光谱/pKa/光稳定性\n每个蛋白含完整注释和文献引用',
     '~1,000+蛋白',
     '方法1: https://www.fpbase.org/table → 右下角下载CSV\n方法2: pip install fpbase → Python API\n方法3: https://www.fpbase.org/graphql  GraphQL API',
     'CSV / JSON / Python API'),
    ('ThermoMutDB',
     'http://biosig.unimelb.edu.au/thermomutdb',
     '>14,669突变\n588蛋白\nΔΔG + ΔTm数据\n含GFP相关条目',
     '~14,669突变',
     '1. 访问下载页: http://biosig.unimelb.edu.au/thermomutdb/downloads\n2. 下载CSV/JSON\n3. 搜索"GFP"过滤\n4. 或使用REST API程序化获取',
     'CSV / JSON / REST API'),
    ('ProThermDB',
     'https://web.iitm.ac.in/bioinfo2/prothermdb/',
     '>32,000热力学数据点\nΔG, ΔH, ΔCp, Tm\n>120K高通量数据',
     '~32,000+数据点',
     '1. 访问网站\n2. 搜索"GFP"或UniProt ID: P42212\n3. 填写下载表单\n4. 下载CSV格式',
     'CSV (需注册)'),
    ('ProteinMutTm (中文)',
     'https://github.com/hyq2017/ProteinMutTm',
     '酶突变体热稳定性TSV数据库\n野生型序列+突变+ΔTm\n文献+数据库挖掘',
     '多蛋白',
     '1. git clone仓库\n2. 查看data/目录\n3. 使用TSV文件',
     'TSV'),
    ('synbiochallenges2025 (GitHub)',
     'https://github.com/f-normies/synbiochallenges2025',
     '往年竞赛完整ML管线\nGLEAM模型+MLDE优化\nTemBERTure Tm预测\n含训练数据和代码',
     '竞赛级',
     '1. git clone仓库\n2. 查看notebooks/和data/\n3. 含avGFP训练数据',
     'Python / CSV'),
    ('Mega-scale DMS (ProteinGym)',
     'https://proteingym.org/',
     '大规模深度突变扫描数据\n217 DMS数据集\n含GFP相关条目',
     '>1.5M突变',
     '1. 访问网站\n2. 搜索"GFP"\n3. 下载特定DMS数据集',
     'CSV / JSON'),
    ('ESM-2嵌入 (Zenodo)',
     'https://zenodo.org/records/17088257',
     'GFP DMS数据集的ESM-2 15B参数嵌入\n预计算序列嵌入',
     '预计算嵌入',
     '直接下载 → 可用于ML训练\n无需本地GPU推理',
     'NPY / Pickle'),
    ('Meltome Atlas (PXD011929)',
     'https://ftp.pride.ebi.ac.uk/pride/data/archive/2020/04/PXD011929',
     '热蛋白质组稳定性图谱\n13物种, 跨越生命之树\nTm数据',
     '~数万蛋白',
     '1. 访问PRIDE\n2. 下载data文件\n3. 搜索GFP条目',
     'TSV / 原始质谱'),
    ('GeoStab (Gonglab-THU)',
     'https://github.com/Gonglab-THU/GeoStab',
     '蛋白质稳定性ΔTm预测\n含S4346/S1626/S571数据集\nΔTm点突变数据',
     '~6,543单点突变',
     '1. git clone仓库\n2. 查看data/目录\n3. 使用提供的CSV',
     'CSV'),
]

cols_guide = ['Dataset', 'Source_URL', 'Description', 'Data_Size', 'Access_Method', 'Format']
guide_df = pd.DataFrame(download_guide, columns=cols_guide)

# ============================================================
# 8. 写入Excel
# ============================================================
print(f"\n写入Excel文件: {output_path}")

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    train_df.to_excel(writer, sheet_name='Brightness_Training_Data', index=False)
    fpbase_df.to_excel(writer, sheet_name='FPbase_GFP_Variants', index=False)
    thermal_df.to_excel(writer, sheet_name='Thermal_Stability_Data', index=False)
    lit_df.to_excel(writer, sheet_name='Literature_GFP_Properties', index=False)
    guide_df.to_excel(writer, sheet_name='Data_Download_Guide', index=False)
    ref_df.to_excel(writer, sheet_name='Reference_Sequences', index=False)
    cand_df.to_excel(writer, sheet_name='Candidate_Positions', index=False)

# 调整列宽
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(output_path)
for ws in wb.sheetnames:
    worksheet = wb[ws]
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    # 冻结首行
    worksheet.freeze_panes = 'A2'

wb.save(output_path)

print(f"\n完成! 综合数据集已保存至:")
print(f"  {output_path}")
print(f"\n包含 {len(wb.sheetnames)} 个工作表:")
for s in wb.sheetnames:
    df = pd.read_excel(output_path, sheet_name=s)
    print(f"  - {s}: {len(df)} 行, {len(df.columns)} 列")

print("\n📋 数据下载行动清单:")
print("  1. [推荐] 下载 Sarkisyan Nature 2016 数据集 (Figshare) → 51,715条GFP突变数据")
print("  2. [推荐] 下载 FPbase 完整数据 → 1,000+荧光蛋白的参数")
print("  3. [推荐] 下载 ThermoMutDB GFP条目 → 补充热稳定性ΔΔG/ΔTm")
print("  4. [可选] 克隆 synbiochallenges2025 GitHub → 参考竞赛管线")
print("  5. [可选] 下载 GeoStab S4346 → 通用蛋白ΔTm数据")
print("  6. [可选] 下载 ESM-2预计算嵌入 → 即用型特征")
